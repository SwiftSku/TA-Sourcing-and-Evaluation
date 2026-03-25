"""Sync temp CSV data into the output Excel file. Run after every CE verdict."""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

import os
_dir = os.path.dirname(os.path.abspath(__file__))
CSV_PATH = os.path.join(_dir, "_OUTPUT--Acct_Mgr.csv")
XLSX_PATH = os.path.join(_dir, "_OUTPUT--Acct_Mgr.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "MAIN"

hf = Font(name='Arial', size=10, bold=True, color='000000')
df = Font(name='Arial', size=10)
ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
da = Alignment(wrap_text=True)

with open(CSV_PATH, 'r') as f:
    reader = csv.reader(f)
    for row_idx, row in enumerate(reader, 1):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = hf if row_idx == 1 else df
            cell.alignment = ha if row_idx == 1 else da

widths = {1:22, 2:30, 3:10, 4:30, 5:22, 6:28, 7:16, 8:8, 9:35,
          10:8, 11:25, 12:8, 13:25, 14:8, 15:25, 16:8, 17:25,
          18:8, 19:25, 20:8, 21:25, 22:8, 23:25, 24:8, 25:25,
          26:8, 27:25, 28:10, 29:10, 30:10, 31:10, 32:10, 33:10,
          34:6, 35:12, 36:40, 37:32}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = 'B2'
ws.auto_filter.ref = f"A1:{get_column_letter(37)}1"

wb.save(XLSX_PATH)
print(f"Synced {row_idx - 1} data rows to {XLSX_PATH}")
