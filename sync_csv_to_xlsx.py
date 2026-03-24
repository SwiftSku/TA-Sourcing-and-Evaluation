"""Sync temp CSV data into the output Excel file. Run after every CE verdict."""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

CSV_PATH = "/sessions/awesome-determined-brown/_temp_rc.csv"
XLSX_PATH = "/sessions/awesome-determined-brown/mnt/TA-ACM/_OUTPUT--AMD_Recruiting_Coord.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "MAIN"

hf = Font(name='Arial', size=10, bold=True, color='000000')
df = Font(name='Arial', size=10)
ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
da = Alignment(wrap_text=True)

with open(CSV_PATH, 'r') as f:
    reader = csv.reader(f)
    for row_idx, row in enumerate(reader, 1):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = hf if row_idx == 1 else df
            cell.alignment = ha if row_idx == 1 else da

widths = {1:22, 2:10, 3:30, 4:22, 5:28, 6:16, 7:8, 8:35,
          9:8, 10:25, 11:8, 12:25, 13:8, 14:25, 15:8, 16:25,
          17:8, 18:25, 19:8, 20:25, 21:8, 22:25, 23:8, 24:25,
          25:8, 26:25, 27:10, 28:10, 29:10, 30:10, 31:10, 32:10,
          33:6, 34:12, 35:40, 36:32, 37:10}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = 'B2'
ws.auto_filter.ref = f"A1:{get_column_letter(37)}1"

wb.save(XLSX_PATH)
print(f"Synced {row_idx - 1} data rows to {XLSX_PATH}")
